
import requests
from bs4 import BeautifulSoup
import pandas as pd
#Convert Excel work
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


LOGIN_URL = "https://cms.must.edu.pk/sfms/checklogin.asp"
REPORT_URL = "https://cms.must.edu.pk/sfms/Reports/SemesterFeeCategory/Fee_Summary_Report.asp"
CREDENTIALS = {
    "username": "FEE",          
    "password": "Eef@tomust"    
}
SEMESTER_VALUE = "90-Spring 2025"  
SCHEME_VALUE = "1"               

def main():
   
    with requests.Session() as session:
       
        print("Attempting login...")
        login_response = session.post(LOGIN_URL, data=CREDENTIALS)
        
        if "Login Failed" in login_response.text or login_response.url == LOGIN_URL:
            print("Login failed. Check credentials or network connection.")
            return
        
        print("Login successful")
        
      
        print("Loading report page...")
        report_response = session.get(REPORT_URL,timeout=60)
        report_soup = BeautifulSoup(report_response.text, 'html.parser')
        
       
        form_data = {
            "SelectSession": SEMESTER_VALUE,
            "SelectScheme": SCHEME_VALUE,
            "Submit": "Show Report"
        }
      
        print(" Fetching fee data...")
        data_response = session.post(REPORT_URL, data=form_data)
        data_soup = BeautifulSoup(data_response.text, 'html.parser')
        
       
        data_table = data_soup.find('table', {'name': 'formtable'})
        
        if not data_table:
            print("Data table not found. Possible issues:")
            print("- Semester/Scheme values need updating")
            print("- Page structure changed")
            with open('error_page.html', 'w', encoding='utf-8') as f:
                f.write(data_response.text)
            print("Saved error page to 'error_page.html'")
            return  
        print("Data table found")
        
        
        headers = [th.get_text(strip=True) for th in data_table.find('tr').find_all('td')]
        print(headers)
      
        science_departments=["Biotechnology","Chemistry","Computer Sciences & Information Technology","Forestry","Mathematics","Physics","Physics (Pallandri Campus)","Statistics","Zoology"]
        engineering_departments=["Civil Engineering","Civil Engineering Technology","Computer Systems Engineering","Electrical Engineering","Electrical Engineering Technology","Mirpur Institute of Technology","Mechanical Engineering","Software Engineering"]
        social_sciences_departments =["Education","Education (Pallandri Campus)","English","Home Economics","Institute of Islamic Studies","International Relations","International Relations (Pallandri Campus)","Law","Mass Communication","Sociology","Tourism and Hospitality"]
        health_medical_departments =["Allied Health Sciences","Human Nutrition & Dietetics","Microbiology","Pharmacy","Physiotherapy"]
        must_businessSchool_departments=["MUST business School"]
        
        faculties={}
        faculty_sciences={}
        faculty_Engineering={}
        faculty_social_sciences={}
        faculty_health_medical={}
        faculty_must_business={}

        for row in data_table.find_all('tr')[1:]:  # Skip header row
            
            if row.find_all('td')[0].get_text(strip=True) in science_departments:
                faculty_sciences[row.find_all('td')[0].get_text(strip=True)]={headers[4]:row.find_all('td')[4].get_text(strip=True),headers[11]:row.find_all('td')[11].get_text(strip=True),headers[12]:row.find_all('td')[12].get_text(strip=True)}
                faculties["Faculty of Natural and Applied Sciences"]=faculty_sciences
            elif row.find_all('td')[0].get_text(strip=True) in engineering_departments:
                faculty_Engineering[row.find_all('td')[0].get_text(strip=True)]={headers[4]:row.find_all('td')[4].get_text(strip=True),headers[11]:row.find_all('td')[11].get_text(strip=True),headers[12]:row.find_all('td')[12].get_text(strip=True)}
                faculties["Faculty of Engineering and Technology"]=faculty_Engineering
            elif row.find_all('td')[0].get_text(strip=True) in social_sciences_departments:
                faculty_social_sciences[row.find_all('td')[0].get_text(strip=True)]={headers[4]:row.find_all('td')[4].get_text(strip=True),headers[11]:row.find_all('td')[11].get_text(strip=True),headers[12]:row.find_all('td')[12].get_text(strip=True)}
                faculties["Faculty of Social Sciences and Humanities"]=faculty_social_sciences
            elif row.find_all('td')[0].get_text(strip=True) in health_medical_departments:
                faculty_health_medical[row.find_all('td')[0].get_text(strip=True)]={headers[4]:row.find_all('td')[4].get_text(strip=True),headers[11]:row.find_all('td')[11].get_text(strip=True),headers[12]:row.find_all('td')[12].get_text(strip=True)}
                faculties["Faculty of Health and Medical Sciences"]=faculty_health_medical
            elif row.find_all('td')[0].get_text(strip=True) in must_businessSchool_departments:
                faculty_must_business[row.find_all('td')[0].get_text(strip=True)]={headers[4]:row.find_all('td')[4].get_text(strip=True),headers[11]:row.find_all('td')[11].get_text(strip=True),headers[12]:row.find_all('td')[12].get_text(strip=True)}
                faculties["Faculty of MBS"]=faculty_must_business
        print(faculties)
        create_outstanding_dues_report(faculties, "outstanding_dues_report.xlsx")
            
def create_outstanding_dues_report(data_dict, output_filename):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Set up the header rows
    ws.merge_cells('B1:F1')
    ws['B1'] = "Outstanding Dues Spring Semester 2025 As Per Student Fee Management Record"
    ws['B1'].font = Font(bold=True)
    ws['B1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = [
        "Sr. No.",
        "Faculty",
        "Department",
        "Collected Dues",
        "Outstanding Dues",
        "No. Of Student Outstanding Dues"
    ]
    ws.append(headers)
    
    # Make headers bold
    for cell in ws[2]:  # Headers are in row 2
        cell.font = Font(bold=True)
    
    # Populate the data
    sr_no = 1
    first_data_row = 3  # Data starts at row 3 (after headers in row 2)
    
    for faculty, departments in data_dict.items():
        # For the first department in each faculty, write the faculty name
        first_dept = True
        
        for dept, values in departments.items():
            row = [
                sr_no,
                faculty if first_dept else "",  # Only show faculty name once
                dept,
                values.get('Fee Collected', ''),
                values.get('Outstanding Dues', ''),
                values.get('No. of Students, Outstanding Dues', '')
            ]
            ws.append(row)
            sr_no += 1
            first_dept = False
    
    # Calculate the last data row (header is row 2, first data row is 3)
    last_data_row = 2 + sr_no  # 2 (header) + number of data rows
    
    # Add the totals row (corrected version)
    total_row_num = last_data_row + 1
    ws.append([
        "Total outstanding Amount",
        "",
        "",
        "",
        f"=SUM(E3:E{last_data_row})",
        f"=SUM(F3:F{last_data_row})"
    ])
    
    # Make totals row bold
    for cell in ws[total_row_num]:
        cell.font = Font(bold=True)
    
    # Save the workbook
    wb.save(output_filename)


if __name__ == "__main__":
    main()
